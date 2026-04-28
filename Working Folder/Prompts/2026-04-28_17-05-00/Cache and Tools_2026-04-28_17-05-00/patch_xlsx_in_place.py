#!/usr/bin/env python3
"""
patch_xlsx_in_place.py — surgical alternative to a full rebuild.

Patches only the 26 corrupted Bellevue cells in
achievement_project/dashboard/G3-G7_Achievement_Data.xlsx
without regenerating the whole workbook. Use this if you've made other
downstream edits to the xlsx and don't want to lose them.

Updates two sheets:
  - "All Demographics" — replaces the 26 corrupted rows
  - "G3-G7 Aggregate" — recomputes the n-weighted Bellevue aggregates
    that depend on the corrupted rows

Run from anywhere:
    python3 patch_xlsx_in_place.py

Requires openpyxl. Install via:
    python3 -m pip install openpyxl --break-system-packages

Created 2026-04-28 as part of the Bellevue AIM/SBAC fix bundle.
"""
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: python3 -m pip install openpyxl --break-system-packages")

# Resolve xlsx path relative to this script's location:
#   Working Folder/patch_xlsx_in_place.py
#   ../achievement_project/dashboard/G3-G7_Achievement_Data.xlsx
HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(HERE)  # Achievement/
XLSX_PATH = os.path.join(PROJECT_ROOT, "achievement_project", "dashboard", "G3-G7_Achievement_Data.xlsx")

if not os.path.exists(XLSX_PATH):
    sys.exit(f"ERROR: xlsx not found at {XLSX_PATH}")

# (year, subject, grade, subgroup) -> (corrected N, corrected pct)
PATCHES = {
    # --- 2025 (14 cells, OSPI dataset h5d9-vgwi)
    ("2025", "Math", "3", "All Students"): (1345, 69.4),
    ("2025", "ELA",  "3", "All Students"): (1333, 65.1),
    ("2025", "ELA",  "3", "SWD"):          (170, 34.7),
    ("2025", "Math", "3", "SWD"):          (170, 40.0),
    ("2025", "Math", "4", "All Students"): (1381, 72.9),
    ("2025", "ELA",  "4", "SWD"):          (175, 40.0),
    ("2025", "Math", "4", "SWD"):          (176, 37.5),
    ("2025", "Math", "5", "All Students"): (1335, 67.9),
    ("2025", "ELA",  "5", "SWD"):          (139, 28.1),
    ("2025", "Math", "6", "All Students"): (1486, 65.3),
    ("2025", "ELA",  "6", "SWD"):          (159, 22.0),
    ("2025", "Math", "6", "SWD"):          (159, 25.2),
    ("2025", "Math", "7", "All Students"): (1502, 65.3),
    ("2025", "Math", "7", "SWD"):          (136, 18.4),
    # --- 2024 (4 cells, OSPI dataset x73g-mrqp)
    ("2024", "Math", "4", "All Students"): (1271, 71.5),
    ("2024", "Math", "4", "SWD"):          (148, 32.4),
    ("2024", "Math", "5", "All Students"): (1347, 67.6),
    ("2024", "Math", "5", "SWD"):          (161, 29.2),
    # --- 2022 (8 cells, OSPI dataset v928-8kke)
    ("2022", "ELA",  "3", "All Students"): (1299, 71.6),
    ("2022", "ELA",  "3", "SWD"):          (120, 31.7),
    ("2022", "Math", "3", "All Students"): (1309, 72.8),
    ("2022", "Math", "3", "SWD"):          (121, 34.7),
    ("2022", "ELA",  "6", "All Students"): (1391, 63.8),
    ("2022", "ELA",  "6", "SWD"):          (140, 15.7),
    ("2022", "Math", "6", "All Students"): (1400, 61.2),
    ("2022", "Math", "6", "SWD"):          (140, 16.4),
}


def main():
    print(f"Opening {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # -------- Sheet 2: All Demographics --------
    if "All Demographics" not in wb.sheetnames:
        sys.exit('ERROR: "All Demographics" sheet not found.')
    ws = wb["All Demographics"]

    # Header row 1: District | State | Year | Subject | Grade | Subgroup | N Tested | % Met or Exceeded
    patched = 0
    for row in ws.iter_rows(min_row=2, max_col=8):
        district = row[0].value
        if district != "Bellevue SD":
            continue
        year, subject, grade, subgroup = row[2].value, row[3].value, row[4].value, row[5].value
        key = (str(year), str(subject), str(grade), str(subgroup))
        if key in PATCHES:
            new_n, new_pct = PATCHES[key]
            row[6].value = new_n
            row[7].value = round(new_pct, 1)
            patched += 1
            print(f"  patched {' '.join(key):40s} -> N={new_n}, %={new_pct}")

    print(f"All Demographics: patched {patched} cells (expected 26).")
    if patched != 26:
        print(f"WARNING: expected 26 patches but applied {patched}. Inspect the workbook.")

    # -------- Sheet 3: G3-G7 Aggregate --------
    # Recompute n-weighted aggregates for Bellevue (district, year, subject, subgroup)
    # using the now-corrected per-grade values from "All Demographics".
    bellevue_cells = defaultdict(list)
    for row in ws.iter_rows(min_row=2, max_col=8):
        if row[0].value != "Bellevue SD":
            continue
        year, subject, grade, subgroup = row[2].value, row[3].value, row[4].value, row[5].value
        n, pct = row[6].value, row[7].value
        if str(grade) not in ("3", "4", "5", "6", "7"):
            continue
        if pct is None:
            continue
        try:
            pct = float(pct)
        except (TypeError, ValueError):
            continue
        n_val = int(n) if isinstance(n, (int, float)) and n is not None else None
        bellevue_cells[(str(year), str(subject), str(subgroup))].append((n_val, pct))

    new_aggs = {}
    for key, items in bellevue_cells.items():
        items_n = [x for x in items if x[0] is not None]
        if items_n and len(items_n) == len(items):
            total_n = sum(n for n, _ in items_n)
            agg_pct = sum(n * p for n, p in items_n) / total_n if total_n else None
            new_aggs[key] = (total_n, round(agg_pct, 1)) if agg_pct is not None else (total_n, None)
        else:
            agg_pct = sum(p for _, p in items) / len(items)
            new_aggs[key] = (None, round(agg_pct, 1))

    if "G3-G7 Aggregate" in wb.sheetnames:
        ws_a = wb["G3-G7 Aggregate"]
        agg_patched = 0
        # Header is on row 3. Columns: District|State|Year|Subject|Subgroup|N Total|% Met (n-weighted)
        for row in ws_a.iter_rows(min_row=4, max_col=7):
            if row[0].value != "Bellevue SD":
                continue
            year, subject, subgroup = row[2].value, row[3].value, row[4].value
            key = (str(year), str(subject), str(subgroup))
            if key in new_aggs:
                n_new, pct_new = new_aggs[key]
                row[5].value = n_new
                row[6].value = pct_new
                agg_patched += 1
        print(f"G3-G7 Aggregate: recomputed {agg_patched} Bellevue rollups.")
    else:
        print('NOTE: "G3-G7 Aggregate" sheet not found — skipping rollup recompute.')

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")
    print(f"Size:  {os.path.getsize(XLSX_PATH)/1024:.0f} KB")


if __name__ == "__main__":
    main()
