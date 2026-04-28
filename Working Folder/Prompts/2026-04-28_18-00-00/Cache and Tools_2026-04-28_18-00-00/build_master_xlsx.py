"""Build comprehensive XLSX with all dashboard data on multiple sheets.

Lives in: Working Folder/Cache and Tools/
Reads:    Working Folder/achievement_project/extracted_data/master_demographics.csv
Writes:   Working Folder/achievement_project/dashboard/G3-G7_Achievement_Data.xlsx
"""
import csv, openpyxl, os, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Portable path resolution via shared helper next to this script.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import MASTER_CSV, DASHBOARD_XLSX as OUT

wb = openpyxl.Workbook()

# ============== Sheet 1: README ==============
ws_r = wb.active
ws_r.title = 'README'
ws_r.column_dimensions['A'].width = 110
ws_r['A1'] = 'G3-G7 Achievement Comparison — Master Data Workbook'
ws_r['A1'].font = Font(bold=True, size=16, name='Arial')
readme_lines = [
    '',
    'Multi-state grade-level achievement comparison with full demographic disaggregation.',
    '8 high-performing suburban districts in five states. Spring 2019 - Spring 2025 + partial 2021.',
    '',
    'CONTENTS:',
    '',
    '  Sheet 2: All Demographics — every (district × year × subject × grade × subgroup) cell, n-tested + % met threshold',
    '  Sheet 3: G3-G7 Aggregate — n-weighted G3-G7 aggregates per (district × year × subject × subgroup)',
    '  Sheet 4: Curriculum Adoptions — Math + ELA programs and adoption years per district',
    '  Sheet 5: State Standards — links to authoritative state ELA/Math learning standards',
    '  Sheet 6: District Profile — enrollment, demographics, summary metrics',
    '  Sheet 7: Methodology — caveats, suppression rules, normalization notes',
    '',
    'CROSS-STATE COMPARABILITY:',
    'Each state uses a different test and a different proficiency threshold:',
    '  - CA (CAASPP / Smarter Balanced): "Met or Exceeded Standard"',
    '  - MI (M-STEP): "Proficient or Advanced"',
    '  - TX (STAAR via TAPR): "Meets Grade Level or Above"',
    '  - NJ (NJSLA): "Met or Exceeded Expectations" (Levels 4-5)',
    '  - WA (Smarter Balanced): "Met Standard" (Levels 3-4)',
    '',
    'Cross-state ABSOLUTE % NOT directly comparable. Within-district trends ARE comparable.',
    '',
    'DATA SOURCES:',
    '  MI: MI School Data CEPI Public Performance files (5 years × Performance Level Achievement)',
    '  CA: CAASPP Public Data Files (statewide research files)',
    '  TX: Texas Academic Performance Reports (TAPR) PDFs',
    '  NJ: NJDOE School Performance Reports + 2024-25 NJSLA bulk Excel files',
    '  WA: OSPI Open Data API (data.wa.gov Socrata datasets)',
    '',
    'Last updated: April 2026 (FY26 partial through April 2026)',
]
for i, line in enumerate(readme_lines, 2):
    ws_r.cell(row=i, column=1, value=line).font = Font(name='Arial')

# ============== Sheet 2: All Demographics ==============
ws_d = wb.create_sheet('All Demographics')
headers = ['District', 'State', 'Year', 'Subject', 'Grade', 'Subgroup', 'N Tested', '% Met or Exceeded']
ws_d.append(headers)
for cell in ws_d[1]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F4E78')
    cell.alignment = Alignment(horizontal='center')

with open(MASTER_CSV) as f:
    reader = csv.DictReader(f)
    for r in reader:
        if not r.get('pct_met') or r['pct_met'] in ('', 'None'): continue
        try: pct = float(r['pct_met'])
        except: continue
        try: n = int(r['tested']) if r.get('tested') and r['tested'] not in ('', 'None') else None
        except: n = None
        ws_d.append([r['district'], r['state'], r['year'], r['subject'], r['grade'], r['subgroup'], n, round(pct, 1)])

ws_d.column_dimensions['A'].width = 28
ws_d.column_dimensions['B'].width = 8
for col in 'CDE': ws_d.column_dimensions[col].width = 8
ws_d.column_dimensions['F'].width = 28
ws_d.column_dimensions['G'].width = 12
ws_d.column_dimensions['H'].width = 18
# Format pct column
for row in ws_d.iter_rows(min_row=2, min_col=8, max_col=8):
    for cell in row: cell.number_format = '0.0"%"'
ws_d.freeze_panes = 'A2'
ws_d.auto_filter.ref = f"A1:H{ws_d.max_row}"

# ============== Sheet 3: G3-G7 Aggregate ==============
ws_a = wb.create_sheet('G3-G7 Aggregate')
ws_a['A1'] = 'G3-G7 Aggregate (n-weighted across grades)'
ws_a['A1'].font = Font(bold=True, size=12, name='Arial')

# Build aggregates from demographics
agg_data = defaultdict(list)
with open(MASTER_CSV) as f:
    for r in csv.DictReader(f):
        if not r.get('pct_met') or r['pct_met'] in ('', 'None'): continue
        try: pct = float(r['pct_met'])
        except: continue
        try: n = int(r['tested']) if r.get('tested') and r['tested'] not in ('', 'None') else None
        except: n = None
        if r['grade'] not in ('3','4','5','6','7'): continue
        key = (r['district'], r['state'], r['year'], r['subject'], r['subgroup'])
        agg_data[key].append((n, pct))

# Compute aggregate
aggs = []
for (d, s, y, sub, sg), items in agg_data.items():
    items_n = [x for x in items if x[0] is not None]
    if items_n and len(items_n) == len(items):
        total_n = sum(n for n, _ in items_n)
        agg_pct = sum(n * p for n, p in items_n) / total_n if total_n else None
        agg_n = total_n
    else:
        agg_pct = sum(p for _, p in items) / len(items)
        agg_n = None
    aggs.append({'d': d, 's': s, 'y': y, 'sub': sub, 'sg': sg, 'n': agg_n, 'p': round(agg_pct, 1)})

ws_a.append([])
header_row_idx = ws_a.max_row + 1
ws_a.append(['District', 'State', 'Year', 'Subject', 'Subgroup', 'N Total', '% Met (n-weighted)'])
for cell in ws_a[header_row_idx]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F4E78')
    cell.alignment = Alignment(horizontal='center')

aggs.sort(key=lambda x: (x['d'], x['sub'], x['sg'], x['y']))
for a in aggs:
    ws_a.append([a['d'], a['s'], a['y'], a['sub'], a['sg'], a['n'], a['p']])

ws_a.column_dimensions['A'].width = 28
ws_a.column_dimensions['B'].width = 8
ws_a.column_dimensions['C'].width = 8
ws_a.column_dimensions['D'].width = 10
ws_a.column_dimensions['E'].width = 28
ws_a.column_dimensions['F'].width = 12
ws_a.column_dimensions['G'].width = 18
for row in ws_a.iter_rows(min_row=header_row_idx+1, min_col=7, max_col=7):
    for cell in row: cell.number_format = '0.0"%"'
ws_a.freeze_panes = f'A{header_row_idx+1}'

# ============== Sheet 4: Curriculum Adoptions ==============
ws_c = wb.create_sheet('Curriculum Adoptions')
ws_c['A1'] = 'District Curriculum Adoptions — Math + ELA Programs and Years'
ws_c['A1'].font = Font(bold=True, size=12, name='Arial')

ws_c.append([])
ws_c.append(['District', 'State', 'K-5 Math', 'K-5 Math Year', 'K-5 ELA', 'K-5 ELA Year', '6-8 Math', '6-8 Math Year', '6-8 ELA', '6-8 ELA Year'])
for cell in ws_c[3]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F4E78')

CURRICULUM = [
    ('Troy SD', 'MI', 'Bridges in Mathematics', '2023', 'Units of Study (Calkins/Heinemann)', 'Fall 2014',
     'Imagine IM (Illustrative Math 6-8)', 'May 2023', 'UoS Writing + balanced literacy', '~2014'),
    ('Palo Alto USD', 'CA', 'Bridges in Mathematics 2nd Ed.', 'May 9, 2017', 'Benchmark Advance/Adelante', 'May 2022',
     'Big Ideas Math (Course 1, 2 Acc., 3)', 'March 28, 2017', 'Core literature anthologies (no packaged ELA)', 'May 11, 2021'),
    ('Milpitas USD', 'CA', 'Math in Focus + Go Math! (HMH)', '2013-14', 'Benchmark Advance', 'year not located',
     'Glencoe Math + Go Math! + Big Ideas (acc.)', '~2014-15', 'StudySync (McGraw Hill)', 'year not located'),
    ('Walnut Valley USD', 'CA', 'Math Expressions (HMH)', '2015', 'Benchmark Advance', 'April 19, 2017',
     'Big Ideas (HMH)', '2015', 'California Collections (HMH)', 'April 19, 2017'),
    ('Dublin USD', 'CA', 'Eureka Math Squared (Great Minds)', '~2023-24', 'Benchmark Advance + Heggerty + Calkins UoS Writing', 'year not located',
     'i-Ready Classroom Math (prior: Glencoe)', '2024-25', 'Into Literature (HMH)', '2023-24'),
    ('Coppell ISD', 'TX', 'Texas Go Math! (HMH)', 'March 27, 2014', 'District-built balanced literacy', 'in-house',
     'Texas Go Math! (HMH) — same K-8 adoption', 'March 27, 2014', 'District-built balanced literacy', 'in-house'),
    ('West Windsor-Plainsboro', 'NJ', 'enVisionMATH (Pearson/Savvas) — K-3', 'year not located', 'District-built Units of Study (Calkins/TCRWP)', 'Framework June 2017',
     'EdGems Math + Big Ideas (Pre-Algebra)', 'Pilot July 26, 2016', 'Same district-built UoS K-8', 'Framework June 2017'),
    ('Bellevue SD', 'WA', 'Illustrative Mathematics K-5', '2021-22', 'ARC Core + UFLI Foundations + Heggerty', '2024-25 (ARC Core)',
     'Illustrative Mathematics 6-8 (Imagine IM)', '2021-22', 'No commercial program — district-built', 'not located'),
]
for row in CURRICULUM:
    ws_c.append(list(row))

ws_c.column_dimensions['A'].width = 26
ws_c.column_dimensions['B'].width = 7
for col in ['C','E','G','I']: ws_c.column_dimensions[col].width = 38
for col in ['D','F','H','J']: ws_c.column_dimensions[col].width = 20

# ============== Sheet 5: State Standards ==============
ws_s = wb.create_sheet('State Standards')
ws_s['A1'] = 'State Math + ELA Learning Standards (Source Documents)'
ws_s['A1'].font = Font(bold=True, size=12, name='Arial')
ws_s.append([])
ws_s.append(['State', 'Subject', 'Standards Document', 'URL'])
for cell in ws_s[3]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F4E78')

STANDARDS = [
    ('Michigan', 'Mathematics', 'MI Mathematics K-12 Academic Standards', 'https://www.michigan.gov/-/media/Project/Websites/mde/Year/2018/01/22/K_12_MI_Math_Standards_REV.pdf'),
    ('Michigan', 'ELA', 'MI English Language Arts K-12 Standards', 'https://www.michigan.gov/-/media/Project/Websites/mde/Year/2018/01/22/K-12_MI_ELA_StandardsREV.pdf'),
    ('California', 'Mathematics', 'CA Common Core Mathematics Framework', 'https://www.cde.ca.gov/ci/ma/cf/'),
    ('California', 'ELA', 'CA Common Core ELA / ELD Framework', 'https://www.cde.ca.gov/ci/rl/cf/elaeldfwchapters.asp'),
    ('Texas', 'Mathematics', 'TEKS — Mathematics', 'https://tea.texas.gov/academics/curriculum-standards/teks/texas-essential-knowledge-and-skills'),
    ('Texas', 'ELA', 'TEKS — ELA and Reading', 'https://tea.texas.gov/academics/subject-areas/english-language-arts-and-reading'),
    ('New Jersey', 'Mathematics', 'NJSLS — Mathematics', 'https://www.nj.gov/education/standards/math/'),
    ('New Jersey', 'ELA', 'NJSLS — ELA', 'https://www.nj.gov/education/standards/ela/'),
    ('Washington', 'Mathematics', 'WA K-12 Mathematics Learning Standards', 'https://ospi.k12.wa.us/student-success/learning-standards-instructional-materials/mathematics'),
    ('Washington', 'ELA', 'WA K-12 ELA Learning Standards', 'https://ospi.k12.wa.us/student-success/learning-standards-instructional-materials/english-language-arts'),
]
for row in STANDARDS:
    ws_s.append(list(row))
ws_s.column_dimensions['A'].width = 14
ws_s.column_dimensions['B'].width = 14
ws_s.column_dimensions['C'].width = 50
ws_s.column_dimensions['D'].width = 70

# ============== Sheet 6: District Profile ==============
ws_p = wb.create_sheet('District Profile')
ws_p['A1'] = 'District Demographics + Coverage Summary'
ws_p['A1'].font = Font(bold=True, size=12, name='Arial')
ws_p.append([])
ws_p.append(['District', 'State', 'State Test', 'Threshold Label', 'Subgroups Available', 'Total Cells'])
for cell in ws_p[3]:
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = PatternFill('solid', start_color='1F4E78')

# Compute coverage from master
dist_subgroups = defaultdict(set)
dist_cells = defaultdict(int)
with open(MASTER_CSV) as f:
    for r in csv.DictReader(f):
        if not r.get('pct_met') or r['pct_met'] in ('', 'None'): continue
        dist_subgroups[r['district']].add(r['subgroup'])
        dist_cells[r['district']] += 1

PROFILES = [
    ('Troy SD', 'MI', 'M-STEP', 'Proficient or Advanced'),
    ('Palo Alto USD', 'CA', 'CAASPP', 'Met or Exceeded Standard'),
    ('Milpitas USD', 'CA', 'CAASPP', 'Met or Exceeded Standard'),
    ('Walnut Valley USD', 'CA', 'CAASPP', 'Met or Exceeded Standard'),
    ('Dublin USD', 'CA', 'CAASPP', 'Met or Exceeded Standard'),
    ('Coppell ISD', 'TX', 'STAAR/TAPR', 'Meets Grade Level or Above'),
    ('West Windsor-Plainsboro', 'NJ', 'NJSLA', 'Met or Exceeded Expectations (L4-5)'),
    ('Bellevue SD', 'WA', 'Smarter Balanced', 'Met Standard (L3-4)'),
]
for d, s, t, th in PROFILES:
    sgs = sorted(dist_subgroups[d])
    ws_p.append([d, s, t, th, len(sgs), dist_cells[d]])

ws_p.column_dimensions['A'].width = 26
ws_p.column_dimensions['B'].width = 7
ws_p.column_dimensions['C'].width = 17
ws_p.column_dimensions['D'].width = 38
ws_p.column_dimensions['E'].width = 22
ws_p.column_dimensions['F'].width = 12

# ============== Sheet 7: Methodology ==============
ws_m = wb.create_sheet('Methodology')
ws_m['A1'] = 'Methodology and Caveats'
ws_m['A1'].font = Font(bold=True, size=14, name='Arial')
ws_m.column_dimensions['A'].width = 110
caveats = [
    '',
    '1. Cross-state absolute % NOT directly comparable. CAASPP "Met or Exceeded Standard," M-STEP "Proficient/Advanced,"',
    '   STAAR "Meets Grade Level or Above," NJSLA Levels 4-5, SBA Level 3+4 are different thresholds against different content.',
    '   Within-district trends ARE comparable.',
    '',
    '2. G3-G7 aggregate is n-weighted using each grade\'s tested-student count, except Coppell ISD',
    '   (TAPR doesn\'t publish per-grade tested counts; uses simple G3-G7 mean).',
    '',
    '3. Spring 2020 was canceled nationwide. Spring 2021 had limited admin in MI/CA/NJ/WA. ',
    '   Troy MI 2021 data exists from MI files (partial) but reflects atypical testing conditions.',
    '   Other districts excluded for 2021.',
    '',
    '4. Subgroup naming normalized: "Black/African American," "Hispanic/Latino," "Asian" labels collapsed to canonical names where state-specific labels differ.',
    '',
    '5. Cells suppressed where N < 10 (FERPA). Lines may have year gaps for small subgroups (e.g., American Indian, Pacific Islander).',
    '',
    '6. NJ 2018-19 SPR report has fewer subgroup rows than later years; race/EL/ECD coverage is thinner for that year.',
    '',
    '7. WW-P Multilingual Learner sub-categories (Current ML / Former ML) only available in 2024-25 bulk Excel files; not in older SPRs.',
    '',
    '8. Bellevue WA 2024-25 G3-G5 originally had field-name issues in OSPI Socrata API extraction; corrected April 2026.',
]
for i, c in enumerate(caveats, 2):
    ws_m.cell(row=i, column=1, value=c).font = Font(name='Arial')

# ============== Save ==============
wb.save(OUT)
import os
print(f"Saved: {OUT}")
print(f"Size: {os.path.getsize(OUT)/1024:.0f} KB")
print(f"Sheets: {wb.sheetnames}")
print(f"All Demographics row count: {ws_d.max_row}")
print(f"G3-G7 Aggregate row count: {ws_a.max_row}")
